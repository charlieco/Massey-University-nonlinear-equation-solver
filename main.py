from thermo_functions import *
from sympy import *
from sympy.matrices import SparseMatrix
from openpyxl import *
from CoolProp.CoolProp import PropsSI
import numpy as np
from scipy import optimize
import time,sys,random,math,copy,os,re
from fluprodia import FluidPropertyDiagram
import PySimpleGUI as sg
from joblib import Parallel, delayed

import timeit

sheet_num = 1

p_lim = [0, 0, 0]
t_lim = [0, 0, 0]
s_lim = [0, 0]
h_lim = [0, 0, 0]
x_lim = [0, 1]
mix_lim = [0, 0]

lim = [p_lim, t_lim, s_lim, h_lim, x_lim]
guess = []

F_list = []
F_list_v = []
F_list_vn = []
F_list_non = []
F_list_non_f =[]

V_list = []
V_list_non = []

guess_non = []
J = []
f_error = False
line_instructions = []
working_list = []

coolant1 = []
coolant2 = []
bn1 = []
bn2 = []
removed = []
solution = []
filename = []
excel_sheets = []


def importexcelsheets():#for reading the sheets that are in the Excel file
    global excel_sheets
    print("excel import get sheets", end=" -- ")
    wb = load_workbook(filename, read_only=True, data_only=True) #loads the file
    excel_sheets = wb.sheetnames
    wb.close()
    print("end")


def importexcel(sheet):#Imports all three columns in spreadsheet
    global coolant, working_list, solution
    print("excel import start", end=" -- ")
    wb = load_workbook(filename, read_only=True, data_only=True) #loads the file
    ws = wb[wb.sheetnames[sheet]] #loads the user selected sheet
    z = 0
    solution = [] #solution holds answers to all the variables
    working_list = [] #holds all of the non translated equations
    #go through the Excel sheet col by col
    for col in ws.iter_rows(min_col=1, max_row=200, max_col=5, values_only=True):
        if z != 0: #skips over the first column which are the headers
            if col[0] != ' ':
                if col[0] != None:
                    working_list.append(col[0].lower())
            if col[1] != None:
                if col[1] != '':
                    if col[1] != ' ':
                        solution.append([col[1].lower(), col[2]])
        z = z + 1
    wb.close()
    print("end")
    trans()


def exportexcel(sheet, tosave): #Save the answers back to the Excel sheet
    print("excel save start", end=" -- ")
    wb = load_workbook(filename)
    ws = wb[wb.sheetnames[sheet]]
    for z in range(len(tosave)):
        ws.cell(z + 2, 2).value = tosave[z][0]
        ws.cell(z + 2, 3).value = tosave[z][1]
    wb.save("eqs.xlsx")
    print('end')


def trans():
    # translation function for translating the ees equations to a format that is recognised by simpy
    global V_list, F_list_v, F_list_vn, F_list, coolant2, removed
    print("trans start", end=" -- ")
    F_list = []  # list of all all of the equations
    V_list = []  # list of all the variables
    F_list_v = []  # list of the variables in each equation
    coolant1 = []  # list containing the variables for coolants
    coolant2 = []  # list containing the corresponding values for the variables in coolant1
    n1 = []  # contains the variable for ees duplicate function
    n2 = []  # contains the value for ees duplicate function
    n3 = []  # contains the starting value and end value for I
    for item in working_list:
        ii = False  # variable for showing that I contained within the equation for when duplicate is in eq
        split_work = item.replace(' ', '').split('=', 1)  # splitting the equation by the equals sign
        if split_work[0] == 'end':  # making sure the end statement does not get added to the list of equations
            pass
        elif split_work[1][0] == "'":
            coolant1.append(sympify(split_work[0].split("$")[0]))
            coolant2.append(split_work[1].split("'")[1].upper())
        elif item.split(' ', 1)[0] == "duplicate":  # for detecting the duplicate function in ees
            n3 = [int(item.split(' ', 1)[1].split(',', 1)[0].split('=', 1)[1]),
                  n2[n1.index(sympify(item.split(' ', 1)[1].split(',', 1)[1]))]]  # for starting value and end value for I
        else:
            if split_work[0][0] == 'n':  # for detecting the value that the duplicate statement iterates to. this value needs to start with n
                n1.append(sympify(split_work[0]))  # contains the variable for ees duplicate function
                n2.append(sympify(split_work[1]))  # contains the starting value and end value for I
            if '[' in split_work[0]:  # for detecting if there is square brackets for the duplicate function on the left side of the equation
                split_b1 = split_work[0].split("[")  # splitting at the square brackets
                split_b2 = [x.split("]") for x in split_b1]  # splitting at the square brackets
                split_b2 = [j for sub in split_b2 for j in sub]  # splitting at the square brackets
                sym = [x.name for x in sympify(split_b2[1]).free_symbols]  # calculating all of the variables in current equation
                if 'i' in sym:  # calculating if I is one of the variables
                    ii = True
                    split_work_f = [split_b2.copy() for x in range(n3[1])]  # creating a separate equation for each iteration of I
                    for y in range(n3[0] - 1, n3[1]):  # iterating through each of the newly created equations
                        split_work_f[y][1] = str(y + 1)  # replacing I with the number of the iteration creating a new variable for each iteration
                        split_work_f[y] = ''.join(split_work_f[y])  # rejoins the split equation
                else:
                    split_b2[1] = str(lambdify(n1, sympify(split_b2[1]), "numpy")(*n2))  # substituting in the variable inside the square brackets
                    split_work_f = ''.join(split_b2)  # rejoins the split equation
            else:
                split_work_f = split_work[0]  # temporary variable to store the equation
            if '[' in split_work[1]: # for detecting if there is square brackets for the duplicate function on the right to side of the equation
                split_b1 = split_work[1].split("[")  # splitting at the square brackets
                split_b2 = [x.split("]") for x in split_b1]  # splitting at the square brackets
                split_b2 = [j for sub in split_b2 for j in sub]  # splitting at the square brackets
                sym = [sympify(split_b2[x]).free_symbols for x in range(len(split_b2)) if((x) % 2) == 1]  # calculating all of the variables in current equation
                sym = [j for sub in sym for j in sub]  # calculating all of the variables in current equation
                if 'i' in sym or ii:  # calculating if I is one of the variables
                    ii = True
                    split_work_b = [split_b2.copy() for x in range(n3[1])]  # creating a separate equation for each iteration of I
                    for y in range(n3[0] - 1, n3[1]):  # iterating through each of the newly created equations
                        for z in range(len(split_b2)):  # iterates through the list of split equation split at Square bracket
                            if (( z) % 2) == 1:  # the values of I on your curry every second number Checks if number is even
                                split_work_b[y][z] = str(lambdify(n1, sympify(split_work_b[y][z]), "numpy")(*n2))  # substitute out any "n" variable
                                split_work_b[y][z] = str(lambdify(sympify('i'), sympify(split_work_b[y][z]), "numpy")(y + 1))  # substitute out I
                        split_work_b[y] = ''.join(split_work_b[y])  # rejoins the split equation
                        try:  # attempts to substitute in ant i values that are outside of the square brackets
                            split_work_b[y] = str(lambdify(sympify('i'), sympify(split_work_b[y]), "numpy")(y + 1))
                        except:
                            pass
                else:  # if i is not in the equation
                    for y in range(len(split_b2)):  # iterates through the equations
                        if ((y) % 2) == 1:  # Checks if number is even
                            split_b2[y] = str(
                                lambdify(n1, sympify(split_b2[y]), "numpy")(*n2))  # substitute out any "n" variable
                    split_work_b = ''.join(split_b2)  # rejoins the split equation
            else:
                split_work_b = split_work[1]  # temporary variable to store the equation
            try:  # attempts to convert equations to sympy equation
                if ii:  # if i is present in the equation
                    for z in range(len(split_work_f)):  # iterate through each equation created because of I
                        sub_work = (sympify(split_work_b[z]) - sympify(split_work_f[z]))  # sympify does the conversion
                        F_list.append(sub_work)  # add the new equation to the list of equations
                else:
                    sub_work = (sympify(split_work_b) - sympify(split_work_f))
                    F_list.append(sub_work)  # add the new equation to the list of equations

            # if equation cannot be converted it means that there is a cool prop / roof prop call within the equation
            except:
                if ii: # if i is present in the equation
                    for y in range(len(split_work_b)): #iterate through the equations with variations of I
                        sub_work = [] #erasers temporary variable
                        worklist = [] #erasers temporary variable
                        test = split_work_b[y].split('(') #splits equation by bracket
                        test1 = (test[1].split(')'))[0].split(',') #splits equation by bracket comma
                        sub_work.append(test[0] + "_") #add the name of the cool prop / roof prop call to the temporary list
                        for z in range(len(test1)): #iterates through the list of the splitequations
                            if z == 0: # the first thing in this list will be the coolant
                                sub_work.append(coolant2[coolant1.index(sympify(test1[0].split('$')[0]))]) #substitute to the coolant valuable for the actual coolant and adds to a list sub_work
                            else:
                                worklist.append(test1[z].replace(' ', '')) #add the variables inside the brackets to the temporary list
                        worklist = sorted(worklist) #sort the work list by alphabetical order
                        for item in worklist:
                            sub_work.append(sympify(item.split('=')[1])) #separates the variables from the equations add them to the temporary list sub_work
                            sub_work[0] = sub_work[0] + item.split('=')[0] #add to the variable to the First type of core cool prop / roof prop call
                        sub_work.append(sympify(split_work_f[y])) #converts variables to sympy equations
                        F_list.append(sub_work) #add the final equation to the list of equations
                else: #performs the same function as the above if component just only performs at once
                    sub_work = [] #erasers temporary variable
                    worklist = [] #erasers temporary variable
                    test = split_work_b.split('(') #splits equation by bracket
                    test1 = (test[1].split(')'))[0].split(',') #splits equation by bracket comma
                    sub_work.append(test[0] + "_") #add the name of the cool prop / roof prop call to the temporary list
                    for z in range(len(test1)): # the first thing in this list will be the coolant
                        if z == 0: # the first thing in this list will be the coolant
                            sub_work.append(coolant2[coolant1.index(sympify(test1[0].split('$')[0]))]) #substitute to the coolant valuable for the actual coolant and adds to a list sub_work
                        else:
                            worklist.append(test1[z].replace(' ', '')) #add the variables inside the brackets to the temporary list
                    worklist = sorted(worklist)  #sort the work list by alphabetical order
                    for item in worklist:
                        sub_work.append(sympify(item.split('=')[1])) #separates the variables from the equations add them to the temporary list sub_work
                        sub_work[0] = sub_work[0] + item.split('=')[0] #add to the variable to the First type of core cool prop / roof prop call
                    sub_work.append(sympify(split_work_f)) #converts variables to sympy equations
                    F_list.append(sub_work) #add the final equation to the list of equations
    # convert the F_list into F_list_v
    for item in F_list: #iterates through all of the items in F_list
        if isinstance(item, list): #Checks if equation is list a list of the variables in each equation
            work1 = []
            for z in range(len(item) - 2): #iterates through the elements of the equation
                if len(item[z + 2].free_symbols) != 0: #checks if element is variable
                    work1.append(list(item[z + 2].free_symbols)[0]) #add variable to temporary list
            F_list_v.append(work1) #add temporary list to mane list
        else:
            F_list_v.append(list(item.free_symbols)) #add variables to list free_symbols all equations from equations from sympy

    V_list = [item for sublist in F_list_v for item in sublist] #reduces dimensions of list
    V_list = list(dict.fromkeys(V_list)) # convert set to dictionary to remove double up variables
    V_list = sympify(V_list)
    removed = []
    # loop removes equations that are one variable that equals another variable variable is substituted into all equations
    for z in range(len(F_list)): #iterates through the elements of F_list
        if len(F_list_v[z]) == 2: #Checks if there are two variables in the equation
            if not isinstance(F_list[z], list): #Checks if equation is not cool prop / refprop equation
                if F_list[z] != 0: #check if equation is not 0
                    fv = sympify(sorted([x.name for x in F_list[z].free_symbols])) #temporary list with equations in it equations are currently strings
                    removed.append([fv[0], fv[1]]) #add the variable that is being removed to a list and the one that is replacing it
                    for y in range(len(F_list)): #iterate through all of the list to substitute in the new variables
                        if isinstance(F_list[y], list): #Checks if equation is not cool prop / refprop equation
                            for z in range(len(F_list[y]) - 2): #iterate through the elements of cool prop / refprop equation
                                F_list[y][z + 2] = F_list[y][z + 2].subs(fv[1], fv[0]) #substitutes in new variable
                        else:
                            F_list[y] = F_list[y].subs(fv[1], fv[0]) #substitutes in new variable
    F_list = [x for x in F_list if x != 0]
    # convert the F_list into F_list_v
    F_list_v=[]
    for item in F_list: #iterates through all of the items in F_list
        if isinstance(item, list): #Checks if equation is list a list of the variables in each equation
            work1 = []
            for z in range(len(item) - 2): #iterates through the elements of the equation
                if len(item[z + 2].free_symbols) != 0: #checks if element is variable
                    work1.append(list(item[z + 2].free_symbols)[0]) #add variable to temporary list
            F_list_v.append(work1) #add temporary list to mane list
        else:
            F_list_v.append(list(item.free_symbols)) #add variables to list free_symbols all equations from equations from sympy

    V_list = [item for sublist in F_list_v for item in sublist] #reduces dimensions of list
    V_list = list(dict.fromkeys(V_list)) # convert set to dictionary to remove double up variables
    V_list = sympify(V_list)
    print(len(F_list),len(F_list_v))
    lims()


def lims():
    global guess, solved, p_lim, t_lim, mix_lim
    print("lims start", end=" -- ")
    for item in coolant2:
        try: #For testing if the fluid is a valid cool proper fluid
            PropsSI("PMIN", item)
            p_lim[0] = PropsSI("PMIN", coolant2[0]) / 100000
            p_lim[2] = PropsSI("Pcrit", coolant2[0]) / 100000
            p_lim[1] = p_lim[2] + 300
            t_lim[0] = PropsSI("TMIN", coolant2[0]) - 273.15
            t_lim[2] = PropsSI("Tcrit", coolant2[0]) - 273.15
            t_lim[1] = t_lim[2] + 300
            h_lim[0] = enthalpy_pt(coolant2[0], [p_lim[1], t_lim[0]])
            h_lim[1] = enthalpy_pt(coolant2[0], [p_lim[0], t_lim[1]])
            h_lim[2] = enthalpy_pt(coolant2[0], [p_lim[2], t_lim[2]])
            s_lim[0] = entropy_pt(coolant2[0], [p_lim[0], t_lim[0]])
            s_lim[1] = entropy_pt(coolant2[0], [p_lim[0], t_lim[1]])
            mix_lim[0] = min([p_lim[0], t_lim[0], h_lim[0]])
            mix_lim[1] = max([p_lim[1], t_lim[1], h_lim[1]])
            guess = []
            solved = []
            for x in range(len(V_list)):
                solved.append(0)
                guess.append(0)
        except:
            print(item," is not a coolprop fluid")
            f_error =True
    print("end ")


def con(x, flist, vlist, na): #Evaluates the list of Equations by substituting values into the equations.
    #tic = time.perf_counter()
    f = []
    for y in range(len(flist)): # iterates through the equations in flist
        if isinstance(flist[y], list): #if equation features function
            c1 = [] #Stores the input variables for the function
            for z in range(len(flist[y]) - 3):
                c1.append(lambdify(vlist, flist[y][z + 2], "numpy")(*x))
            try:
                cool_c = globals()[flist[y][0]](flist[y][1], c1)
                f.append(cool_c - flist[y][len(flist[y]) - 1])  # Add the solution to a temporary list
            except:
                print("no function matching",flist[y][0],flist[y])
                f_error =True
        else:
            f.append(flist[y]) #Add a Standard equation to a temporary list
    #tic1 = time.perf_counter()
    out = lambdify(vlist, f, "numpy")(*x) #Substitutes in values for all of the variables
    out = [abs(ele) for ele in out] #Takes the absolute value of all of the answers To convert the problem from a roots problem to a minimum problem
    #print("c=",time.perf_counter() - tic,time.perf_counter() - tic1)
    return out



def con_single(x, flist, vlist): #Evaluates a single equation by substituting values into the equations.
    if isinstance(flist, list): #if equation features function
        c1=[] #Stores the input variables for the function
        for z in range(len(flist)-3):
            c1.append(lambdify(vlist, flist[z+2], "numpy")(*x))
        try:

            cool_c = eval(flist[0] + "(flist[1],c1)") #Runs the function in the equation
            sub_work = cool_c - flist[len(flist) - 1] #Add the solution to a temporary list
        except:
            print("no function matching", flist[0], flist)
            f_error = True
    else:
        sub_work = flist #Add a Standard equation to a temporary list
    return abs(lambdify(vlist, sub_work, "numpy")(*x)) #Substitutes in values for all of the variables And takes the absolute


def conj(x, flist, vlist, j): #Evaluates The jacobian Matrix by substituting values into the equations.
    #tic = time.perf_counter()
    jj = np.array(copy.deepcopy(j))
    for y in range(len(flist)):# iterates through the equations in flist
        if isinstance(flist[y], list):#if equation features function
            vl = [] #List of the variables in the function
            xl = [] # Lists of the values for those variables
            for z in range(len(flist) - 3): # iterates through Variables in function
                try:
                    flist[y][z + 2].name  # Checks if it is a variable or a number
                    vl.append(flist[y][z + 2])
                    xl.append(x[vlist.index(flist[y][z + 2])])
                except:
                    pass
            temp1 = optimize.approx_fprime(xl, con_single, 0.0001, flist[y], vl) #Estimates the gradient for the function
            for z in range(len(vl)):
                jj[y][vlist.index(vl[z])] = float(temp1[z]) #Adds the answer from the estimate into its corresponding column in the Matrix
    out = Parallel(n_jobs=-1)(delayed(conj_lam)(vlist, j, x) for j in jj) # Send each role of a function Evaluate and sub in values in parallel
    #print("j=",time.perf_counter() - tic)
    return out


def conj_lam(vlist, j, x): #For evaluating the matrix in parallel
    out = lambdify(vlist, SparseMatrix(j), "numpy")(*x)
    out = [i if not math.isnan(i) else 9999999 for i in np.squeeze(out)]  #checks for divide by 0 error
    return out


def algebra_solver(): #Attempts to solve each equation algebraically
    print("start algebra solver", end=" -- ")
    solved_sum = [0, 1] #Stores how many equations were solved on each iteration of the loop
    y = 0
    while solved_sum[y] != solved_sum[y + 1]: #Checks if the last iteration did not solve any more than the previous
        y = y + 1 #iterates through the variable y
        for x in range(len(F_list_v)): #iterates through the list F_list_v
            z = 0
            for i in F_list_v[x]: #iterates through the sublist of F_list_v
                z = z + solved[V_list.index(i)] #Checks if the variable has been solved
                if solved[V_list.index(i)] == 0: #Records what variable is not solved
                    loc = V_list.index(i)
            if len(F_list_v[x]) - z == 1: #Checks if only one variable is not solved
                if isinstance(F_list[x], list):#if equation features function
                    out = con_single(guess, F_list[x], V_list) #Calculates the answer to the equation with the current guess
                    guess1 = guess.copy() #creates a copy of the list of guess
                    guess2 = guess.copy() #creates a copy of the list of guess
                    guess1[loc] = guess1[loc] + out  #adds the answer of the equation to the guess
                    guess2[loc] = guess2[loc] - out  #Subtracts the answer of the equation to the guess
                    if round(con_single(guess1, F_list[x], V_list), 4) == 0: #Checks if the new Guess makes the equation equals 0
                        guess[loc] = guess1[loc]
                        solved[loc] = 1
                    elif round(con_single(guess2, F_list[x], V_list), 4) == 0: #Checks if the new Guess makes the equation equals 0
                        guess[loc] = guess2[loc]
                        solved[loc] = 1
                    else: #Happens if the Solver is not working correctly
                        print("fluid property not solving")
                else:
                    solve_1 = list(solve(F_list[x], V_list[loc])) #Rearrange is the equation algebraically
                    guess[loc] =lambdify(V_list, solve_1[0], "numpy")(*guess) # substitutes in the solved equations
                    solved[loc] = 1 #Sets that equation to be solved
        solved_sum.append(sum(solved)) #Add the total number of solved equations
    print(solved_sum, " end")


def jacobian(): #jacobian matrix is a two dimensional Matrix full of the gradients Each row is the gradient for every variable in the system for one equation
    F_list_w = copy.deepcopy(F_list_non)
    V_list_w = copy.deepcopy(V_list_non)
    print("jacobian start", end=" -- ")
    tic = time.perf_counter()
    jj = np.zeros((len(F_list_w), len(V_list_w)), dtype=object)
    for y in range(len(F_list_w)):
        for z in range(len(V_list_w)):
            if isinstance(F_list_w[y], list): #if equation features function
                for w in range(len(F_list_w[y]) - 2):  # iterates through Variables in function
                    if F_list_w[y][w + 2] == V_list_w[z]:
                        jj[y][z] = F_list_w[y].copy()   #Add the full equation to Matrix to be estimated later on Can not be calculated with function included
            else:
                jj[y][z] = diff((F_list_w[y] ** 2) ** 0.5, V_list_w[z]) #Derivative calculated using the difference in function insympy
    print("end time = ", time.perf_counter() - tic)
    return jj


def non_linear_solver(tol): #
    print("start nonlinear solver")
    tic = time.perf_counter()
    guess_non = []
    bn1 = []
    bn2 = []
    for x in range(len(V_list_non)): #iterates through the equations in V_list_non
        if V_list_non[x].name[0] == 't':  #detects variables starting with T assumed to be temp
            guess_non.append(random.uniform(29, 31)) #Set the temperature to approximately room temperature
            bn1.append(t_lim[0])    #Sets the upper and lower limits
            bn2.append(t_lim[1])
        elif V_list_non[x].name[0] == 'p':
            guess_non.append(random.uniform(9.9, 10.1))
            bn1.append(p_lim[0])
            bn2.append(p_lim[1])
        elif V_list_non[x].name[0] == 'h':
            guess_non.append(random.uniform(h_lim[2] - 10, h_lim[2] + 10)) #Set the enthalpy A random number between the Max and min
            bn1.append(h_lim[0])
            bn2.append(h_lim[1])
        elif V_list_non[x].name[0] == 's':
            guess_non.append(random.uniform(s_lim[0], s_lim[1])) #Set the enthalpy A random number between the Max and min
            bn1.append(s_lim[0])
            bn2.append(s_lim[1])
        else:
            bn1.append(mix_lim[0])
            bn2.append(mix_lim[1])
            guess_non.append(random.uniform(1, 10))
    print("\n\n")
    # The function call for the solver
    sol = optimize.least_squares(con, guess_non, jac=conj, bounds=(bn1, bn2), args=(F_list_non, V_list_non, J),
                                 verbose=2, ftol=tol, xtol=tol, gtol=tol, loss='soft_l1')
    print("\n\ndone\n\n")
    # Evaluates the after for the results from the solver
    after = [abs(ele) for ele in con(sol.x, F_list_non, V_list_non, J)]
    print("\n\n")
    for z in range(len(after)):
        print(round(after[z], 4), ", ", F_list_non[z]) #Displays how well each equation has been  solved
    print("\n\n")
    sol_out = []
    for z in range(len(sol.x)):
        sol_out.append([V_list_non[z].name, round(sol.x[z], 4)]) #Add the variables and their answers to a list for output
    print("end time = ", time.perf_counter() - tic)
    return sol_out


def diagram_draw():
    p1 = [] #Ones list
    h1 = []
    p10 = [] #10s list
    h10 = []
    for x in range(len(solution)):
        if solution[x][0][0] == 'p':    #Detects all variable starting with p
            tempnum= re.sub("[^0-9]", "", solution[x][0])
            if len(solution[x][0])-len(tempnum)==1:
                if len(tempnum) == 1:
                    p1.append([solution[x][0], solution[x][1], int(tempnum)])
                elif len(tempnum) == 2:
                    p10.append([solution[x][0], solution[x][1], int(tempnum)])
        elif solution[x][0][0] == 'h':  #Detects all variable starting with h
            tempnum = re.sub("[^0-9]", "", solution[x][0])
            if len(solution[x][0])-len(tempnum)==1:
                if len(tempnum) == 1:
                    h1.append([solution[x][0], solution[x][1], int(tempnum)])
                elif len(tempnum) == 2:
                    h10.append([solution[x][0], solution[x][1], int(tempnum)])
    p1.extend(p10)
    h1.extend(h10)
    work_p = []
    work_h = []
    if len(p1) == max(len(p1), len(h1)):
        xy1 = np.zeros((h1[len(h1) - 1][2], 3))
    else:
        xy1 = np.zeros((h1[len(h1) - 1][2], 3))
    for z in range(len(p1)):
        xy1[p1[z][2] - 1][1] = p1[z][1]
        work_p.append(p1[z][1])
    for z in range(len(h1)):
        xy1[h1[z][2] - 1][0] = h1[z][1]
        work_h.append(h1[z][1])
    for z in range(len(xy1)):
        xy1[z][2] = z + 1
    for z in range(len(xy1)):
        if xy1[z][0] == 0:
            xy1[z][0] = xy1[z - 1][0]
        if xy1[z][1] == 0:
            xy1[z][1] = xy1[z - 1][1]
    xy1 = list(xy1)
    xy1.append(xy1[0])
    Point_coor = []
    Point_work = []
    Point_exclude = []
    if len(line_instructions[0]) == 4:
        for z in range(len(line_instructions)):
            Point_work.append(list(xy1[int(line_instructions[z][0]) - 1]))
            y = int(line_instructions[z][1])
            while y <= int(line_instructions[z][2]):
                Point_exclude.append(y)
                Point_work.append(list(xy1[y - 1]))
                y = y + 1
            Point_work.append(list(xy1[int(line_instructions[z][3]) - 1]))
            Point_coor.append(Point_work)
            Point_work = []
    for z in range(len(xy1)):
        if xy1[z][2] not in Point_exclude:
            Point_work.append(list(xy1[z]))
    Point_coor.append(Point_work)
    xy_lim = [[max(work_h), max(work_p)], [min(work_h), min(work_p)],
              [(max(work_h) - min(work_h)) / 10, (max(work_p) - min(work_p)) / 10]]
    diagram = FluidPropertyDiagram(coolant2[0])
    diagram.set_unit_system(T='°C', h='kJ/kg', p='bar')
    diagram.set_limits(x_min=xy_lim[1][0] - xy_lim[2][0], x_max=xy_lim[0][0] + xy_lim[2][0],
                       y_min=xy_lim[1][1] - xy_lim[2][1], y_max=xy_lim[0][1] + xy_lim[2][1])
    diagram.calc_isolines()
    diagram.draw_isolines('logph')
    # Loop for drawing all of the points and lines on the diagram
    for z in range(len(Point_coor)):
        for x in range(len(Point_coor[z]) - 1):
            diagram.ax.plot([Point_coor[z][x][0], Point_coor[z][x + 1][0]],
                            [Point_coor[z][x][1], Point_coor[z][x + 1][1]], color='#ff0000')
            diagram.ax.scatter(Point_coor[z][x][0], Point_coor[z][x][1], color='#ff0000')
            diagram.ax.annotate(int(Point_coor[z][x][2]), (Point_coor[z][x][0], Point_coor[z][x][1]), xytext=(5, 5),
                                textcoords='offset pixels', color='#ff0000')
    diagram.save('logph.png', dpi=90)


def solver(): # the main Solver function that is run from the UI button
    global J
    algebra_solver() #use is algebra to solve as many equations as possible
    solved_list = []
    for x in range(len(V_list)):
        if solved[x] == 1: #if equations are solved
            solved_list.append([V_list[x].name, round(guess[x], 4)]) # adds all of the  solved variables and their answers to a list
    if len(V_list) == sum(solved): #Checks if all variables have been solved
        print("all equations have been solved algebraically")
    else:
        test1 = copy.deepcopy(F_list)
        for z in range(len(V_list)): #substitutes in all of the answers to the solved variables
            if solved[z] == 1:
                for y in range(len(F_list)):
                    if isinstance(test1[y], list):
                        for w in range(len(test1[y])-2): # substitutes in solved variables in to thermodynamic functions
                            test1[y][w+2] = test1[y][w+2].subs(V_list[z], guess[z])
                    else: # substitutes in solved variables in to standard equations
                        test1[y] = test1[y].subs(V_list[z], guess[z])
            else: #add the variables that have not been solved to the list of variables for the main solver
                V_list_non.append(V_list[z])
        for z in range(len(F_list)):
            if isinstance(test1[z], list):
                try: #Checks if the equations for the thermodynamic functions have already been solved
                    float(test1[z][len(test1[z])-1])
                except: #only add equations that haven't been solved yet
                    F_list_non.append(test1[z])
            elif test1[z] != 0: #Checks if the equations have already been solved
                F_list_non.append(test1[z])

        J = jacobian() #generates the jacobian Matrix

        x_out = non_linear_solver(1e-12) #runs the nonlinear solver

        for x in range(len(x_out)):
            solved_list.append([x_out[x][0], x_out[x][1]])
    print(removed)
    for z in range(len(removed)):
        for y in range(len(solved_list)):
            if removed[z][0].name == solved_list[y][0]:
                solved_list.append([removed[z][1].name, solved_list[y][1]])
    sorted_out = []
    test_dictionary = dict(solved_list)
    for i in sorted(test_dictionary):
        sorted_out.append([i, test_dictionary[i]])
    print(sorted_out)
    print(sheet_num)
    exportexcel(sheet_num, sorted_out)

#Interface layout instructions
image_viewer_column = [
    [sg.Image(key="-IMAGE-")],
]
file_list_column = [
    [sg.Button('draw', key='BUTTON_draw', size=(30, 1)), sg.Button('solve', key='BUTTON_solve', size=(30, 1))],
    [sg.Text("Select Excel document", size=(40, 1), key="TOUT1")],
    [
        sg.Text("Excel document", size=(15, 1)),
        sg.In(enable_events=True, key="-file-", size=(40, 1))
        , sg.FileBrowse(size=(10, 1))
    ],
    [
        sg.Text("Sheet selection", size=(15, 1)),
        sg.Listbox(excel_sheets, size=(52, len(excel_sheets)), key='-sheet-')
    ],
    [sg.Input(key='line1')],
    [sg.Input(key='line2')],
    [sg.Input(key='line3')],
    [sg.Input(key='line4')],
    #[sg.Output(size=(71,10), key='-OUTPUT-')]
]

layout = [
    [
        sg.Column(file_list_column),
        sg.VSeperator(),
        sg.Column(image_viewer_column),
    ]
]


def main():
    global sheet_num, filename, line_instructions

    window = sg.Window("Image Viewer", layout).Finalize()
    while True:
        event, values = window.read()
        if event == "-file-": #When file has been selected
            filename = values["-file-"] #Variable stores the Excel sheet location
            importexcelsheets()
            window["-sheet-"].update(values=excel_sheets) #Displays the Sheeps and the Excel document
            window["TOUT1"].update('Select Excel Sheet')   # updates the display
        if event == "BUTTON_draw": #Draw button pressed
            if values['-sheet-']:
                line_instructions = [] #The redrawing instructions list
                line_instructions.append(values['line1'].split(","))
                line_instructions.append(values['line2'].split(","))
                line_instructions.append(values['line3'].split(","))
                line_instructions.append(values['line4'].split(","))
                sheet_num = excel_sheets.index(values['-sheet-'][0])
                importexcel(sheet_num)
                diagram_draw()
                window["-IMAGE-"].update(filename='logph.png') # updates the display
            elif filename == []:
                window["TOUT1"].update('No Excel file selected') # updates the display
            else:
                window["TOUT1"].update('No sheet selected') # updates the display
        if event == 'BUTTON_solve': #solve button pressed
            if values['-sheet-']:
                sheet_num = excel_sheets.index(values['-sheet-'][0])
                importexcel(sheet_num)
                solver()
            elif filename == []:
                window["TOUT1"].update('No Excel file selected') # updates the display
            else:
                window["TOUT1"].update('No sheet selected') # updates the display
        if event == "Exit" or event == sg.WIN_CLOSED:
            break

    window.close()


if __name__ == '__main__':
    main()
